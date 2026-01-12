"""
Lookup Table GUI Tool
Loads sensor comparison data files and creates lookup tables for Python-side correction.
"""

import os
import csv
import json
import datetime
import tkinter as tk
import openpyxl
from tkinter import ttk, filedialog, messagebox
import numpy as np
import matplotlib.pyplot as plt
from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg, NavigationToolbar2Tk


class LookupTable:
    """Lookup table class for distance correction"""
    
    def __init__(self, name="Untitled"):
        self.name = name
        self.positions = []  # Reference positions (e.g., string pot readings)
        self.distances = []  # Corresponding sensor distances
        self.created_date = datetime.datetime.now().isoformat()
        self.source_files = []
        self.metadata = {}
    
    def add_data(self, positions, distances):
        """Add data points to the lookup table"""
        self.positions.extend(positions)
        self.distances.extend(distances)
    
    def compile(self, bin_size=1.0, method='average', smooth=False, smooth_method='moving_average', smooth_window=5):
        """
        Compile the lookup table by binning and averaging data points.
        
        Args:
            bin_size: Size of position bins in mm
            method: 'average', 'median', or 'linear_fit'
            smooth: Whether to apply smoothing to the compiled data
            smooth_method: 'moving_average', 'savgol', or 'gaussian'
            smooth_window: Window size for smoothing (must be odd for savgol)
        """
        if not self.positions or not self.distances:
            return False
        
        # Create pairs and sort by position
        data_pairs = list(zip(self.positions, self.distances))
        data_pairs.sort(key=lambda x: x[0])
        
        # Bin the data
        binned_data = {}
        for pos, dist in data_pairs:
            bin_key = round(pos / bin_size) * bin_size
            if bin_key not in binned_data:
                binned_data[bin_key] = []
            binned_data[bin_key].append(dist)
        
        # Calculate representative value for each bin
        self.compiled_positions = []
        self.compiled_distances = []
        self.compiled_std = []
        self.compiled_count = []
        
        for pos in sorted(binned_data.keys()):
            values = binned_data[pos]
            self.compiled_positions.append(pos)
            self.compiled_count.append(len(values))
            self.compiled_std.append(np.std(values) if len(values) > 1 else 0)
            
            if method == 'average':
                self.compiled_distances.append(np.mean(values))
            elif method == 'median':
                self.compiled_distances.append(np.median(values))
            else:
                self.compiled_distances.append(np.mean(values))
        
        self.metadata['bin_size'] = bin_size
        self.metadata['method'] = method
        self.metadata['compiled_date'] = datetime.datetime.now().isoformat()
        
        # Apply smoothing if requested
        if smooth and len(self.compiled_distances) > smooth_window:
            self.apply_smoothing(smooth_method, smooth_window)
            self.metadata['smoothed'] = True
            self.metadata['smooth_method'] = smooth_method
            self.metadata['smooth_window'] = smooth_window
        else:
            self.metadata['smoothed'] = False
        
        return True
    
    def apply_smoothing(self, method='moving_average', window=5):
        """
        Apply smoothing to the compiled lookup table data.
        
        Args:
            method: 'moving_average', 'savgol', or 'gaussian'
            window: Window size for smoothing
        """
        if method == 'moving_average':
            self._smooth_moving_average(window)
        elif method == 'savgol':
            self._smooth_savgol(window)
        elif method == 'gaussian':
            self._smooth_gaussian(window)
    
    def _smooth_moving_average(self, window):
        """Apply moving average smoothing"""
        if window < 2:
            return
        
        smoothed = []
        half_window = window // 2
        
        for i in range(len(self.compiled_distances)):
            start_idx = max(0, i - half_window)
            end_idx = min(len(self.compiled_distances), i + half_window + 1)
            window_values = self.compiled_distances[start_idx:end_idx]
            smoothed.append(np.mean(window_values))
        
        self.compiled_distances = smoothed
    
    def _smooth_savgol(self, window):
        """Apply Savitzky-Golay filter smoothing"""
        try:
            from scipy.signal import savgol_filter
            # Ensure window is odd and less than data length
            if window % 2 == 0:
                window += 1
            window = min(window, len(self.compiled_distances))
            if window < 3:
                return
            
            # Use polynomial order 2 (or window-1 if window is small)
            polyorder = min(2, window - 1)
            self.compiled_distances = savgol_filter(self.compiled_distances, window, polyorder).tolist()
        except ImportError:
            # Fallback to moving average if scipy not available
            self._smooth_moving_average(window)
    
    def _smooth_gaussian(self, window):
        """Apply Gaussian smoothing"""
        from scipy.ndimage import gaussian_filter1d
        try:
            sigma = window / 3.0  # Standard deviation
            self.compiled_distances = gaussian_filter1d(self.compiled_distances, sigma).tolist()
        except ImportError:
            # Fallback to moving average if scipy not available
            self._smooth_moving_average(window)
    
    def lookup(self, position):
        """
        Look up the corrected distance for a given position using linear interpolation.
        """
        if not hasattr(self, 'compiled_positions') or not self.compiled_positions:
            return None
        
        positions = self.compiled_positions
        distances = self.compiled_distances
        
        # Handle edge cases
        if position <= positions[0]:
            return distances[0]
        if position >= positions[-1]:
            return distances[-1]
        
        # Find interpolation points
        for i in range(len(positions) - 1):
            if positions[i] <= position <= positions[i + 1]:
                x1, y1 = positions[i], distances[i]
                x2, y2 = positions[i + 1], distances[i + 1]
                
                # Linear interpolation
                return y1 + (y2 - y1) * (position - x1) / (x2 - x1)
        
        return None
    
    def reverse_lookup(self, distance):
        """
        Reverse lookup: get position for a given distance.
        """
        if not hasattr(self, 'compiled_positions') or not self.compiled_positions:
            return None
        
        positions = self.compiled_positions
        distances = self.compiled_distances
        
        # Handle edge cases
        if distance <= min(distances):
            idx = distances.index(min(distances))
            return positions[idx]
        if distance >= max(distances):
            idx = distances.index(max(distances))
            return positions[idx]
        
        # Find interpolation points (assuming monotonic relationship)
        for i in range(len(distances) - 1):
            d1, d2 = distances[i], distances[i + 1]
            if (d1 <= distance <= d2) or (d2 <= distance <= d1):
                p1, p2 = positions[i], positions[i + 1]
                
                # Linear interpolation
                if d2 != d1:
                    return p1 + (p2 - p1) * (distance - d1) / (d2 - d1)
        
        return None
    
    def get_correction(self, sensor_distance):
        """
        Get the correction offset for a given sensor distance.
        Returns the difference between the true position and the sensor reading.
        """
        true_position = self.reverse_lookup(sensor_distance)
        if true_position is not None:
            return true_position - sensor_distance
        return 0
    
    def get_interpolated_offset(self, sensor_distance):
        """
        Get interpolated offset by directly interpolating the error curve.
        This builds an error lookup table and interpolates it.
        """
        if not hasattr(self, 'compiled_positions') or not self.compiled_positions:
            return 0
        
        # Build error curve: error = distance - position
        errors = [self.compiled_distances[i] - self.compiled_positions[i] 
                  for i in range(len(self.compiled_positions))]
        distances = self.compiled_distances
        
        # Handle edge cases
        if sensor_distance <= min(distances):
            idx = distances.index(min(distances))
            return errors[idx]
        if sensor_distance >= max(distances):
            idx = distances.index(max(distances))
            return errors[idx]
        
        # Linear interpolation in error space
        for i in range(len(distances) - 1):
            d1, d2 = distances[i], distances[i + 1]
            if (d1 <= sensor_distance <= d2) or (d2 <= sensor_distance <= d1):
                e1, e2 = errors[i], errors[i + 1]
                if d2 != d1:
                    interpolated_error = e1 + (e2 - e1) * (sensor_distance - d1) / (d2 - d1)
                    return interpolated_error
        
        return 0
    
    def _apply_spline_correction(self, sensor_distance):
        """
        Apply correction using cubic spline interpolation.
        Provides smoother interpolation than linear.
        """
        if not hasattr(self, 'compiled_positions') or not self.compiled_positions:
            return None
        
        try:
            from scipy.interpolate import CubicSpline
            
            distances = self.compiled_distances
            positions = self.compiled_positions
            
            # Handle edge cases
            if sensor_distance < min(distances) or sensor_distance > max(distances):
                # Fall back to reverse lookup for out of range
                return self.reverse_lookup(sensor_distance)
            
            # Create cubic spline (distance -> position)
            cs = CubicSpline(distances, positions)
            return float(cs(sensor_distance))
            
        except ImportError:
            # Fall back to reverse lookup if scipy not available
            return self.reverse_lookup(sensor_distance)
    
    def _apply_polynomial_correction(self, sensor_distance, degree=3):
        """
        Apply correction using polynomial fit to the error curve.
        """
        if not hasattr(self, 'compiled_positions') or not self.compiled_positions:
            return None
        
        # Build error curve
        errors = [self.compiled_distances[i] - self.compiled_positions[i] 
                  for i in range(len(self.compiled_positions))]
        distances = self.compiled_distances
        
        # Fit polynomial to error vs distance
        degree = min(degree, len(distances) - 1)  # Can't exceed data points - 1
        coeffs = np.polyfit(distances, errors, degree)
        
        # Evaluate polynomial at sensor_distance
        error = np.polyval(coeffs, sensor_distance)
        
        # Corrected position = sensor_distance - error
        return sensor_distance - error
    
    def _apply_nearest_neighbor(self, sensor_distance):
        """
        Apply correction using nearest neighbor (no interpolation).
        Simply finds the closest LUT entry.
        """
        if not hasattr(self, 'compiled_positions') or not self.compiled_positions:
            return None
        
        distances = self.compiled_distances
        positions = self.compiled_positions
        
        # Find nearest distance
        min_diff = float('inf')
        nearest_idx = 0
        
        for i, dist in enumerate(distances):
            diff = abs(dist - sensor_distance)
            if diff < min_diff:
                min_diff = diff
                nearest_idx = i
        
        return positions[nearest_idx]
    
    def _apply_weighted_average(self, sensor_distance, k=3, power=1.0):
        """
        Apply correction using weighted average of k nearest neighbors.
        Weights are based on inverse distance raised to a power.
        
        Args:
            sensor_distance: Raw sensor reading
            k: Number of nearest neighbors to use
            power: Power for inverse distance weighting (higher = more weight to closer points)
        """
        if not hasattr(self, 'compiled_positions') or not self.compiled_positions:
            return None
        
        distances = self.compiled_distances
        positions = self.compiled_positions
        
        # Calculate distances to all LUT points
        diffs = [(abs(dist - sensor_distance), i) for i, dist in enumerate(distances)]
        diffs.sort()  # Sort by distance
        
        # Take k nearest neighbors
        k = min(k, len(diffs))
        nearest = diffs[:k]
        
        # If exact match, return it
        if nearest[0][0] < 1e-6:
            return positions[nearest[0][1]]
        
        # Weighted average using inverse distance with power
        total_weight = 0
        weighted_sum = 0
        
        for diff, idx in nearest:
            weight = 1.0 / ((diff + 1e-6) ** power)  # Power parameter for weighting
            weighted_sum += positions[idx] * weight
            total_weight += weight
        
        return weighted_sum / total_weight if total_weight > 0 else positions[nearest[0][1]]
    
    def apply_correction(self, sensor_distance, method='reverse_lookup', **kwargs):
        """
        Apply correction to a sensor distance reading using specified method.
        
        Args:
            sensor_distance: Raw sensor reading in mm
            method: Correction method to use:
                - 'reverse_lookup': Find true position by reverse lookup (default, linear)
                - 'error_subtraction': Calculate and subtract interpolated error (linear)
                - 'spline': Cubic spline interpolation (smooth, requires scipy)
                - 'polynomial': Polynomial fit to error curve (smooth, degree 3)
                - 'nearest': Nearest neighbor (no interpolation, simple)
                - 'weighted': Weighted average of k nearest neighbors (smooth)
            **kwargs: Method-specific parameters:
                - degree: Polynomial degree (for 'polynomial', default=3)
                - k: Number of neighbors (for 'weighted', default=3)
                - power: Distance weighting power (for 'weighted', default=1.0)
        
        Returns:
            Corrected position in mm
        """
        if method == 'reverse_lookup':
            # Use reverse lookup to find true position
            return self.reverse_lookup(sensor_distance)
        
        elif method in ['error_subtraction', 'offset_correction']:
            # Calculate interpolated error and subtract from reading
            offset = self.get_interpolated_offset(sensor_distance)
            return sensor_distance - offset
        
        elif method == 'spline':
            # Cubic spline interpolation
            return self._apply_spline_correction(sensor_distance)
        
        elif method == 'polynomial':
            # Polynomial fit to error curve
            degree = kwargs.get('degree', 3)
            return self._apply_polynomial_correction(sensor_distance, degree=degree)
        
        elif method == 'nearest':
            # Nearest neighbor (no interpolation)
            return self._apply_nearest_neighbor(sensor_distance)
        
        elif method == 'weighted':
            # Weighted average of k nearest neighbors
            k = kwargs.get('k', 3)
            power = kwargs.get('power', 1.0)
            return self._apply_weighted_average(sensor_distance, k=k, power=power)
        
        else:
            # Default to reverse lookup
            return self.reverse_lookup(sensor_distance)
    
    def to_dict(self):
        """Convert lookup table to dictionary for saving"""
        data = {
            'name': self.name,
            'created_date': self.created_date,
            'source_files': self.source_files,
            'metadata': self.metadata,
            'raw_positions': self.positions,
            'raw_distances': self.distances,
        }
        
        if hasattr(self, 'compiled_positions'):
            data['compiled_positions'] = self.compiled_positions
            data['compiled_distances'] = self.compiled_distances
            data['compiled_std'] = self.compiled_std
            data['compiled_count'] = self.compiled_count
        
        return data
    
    @classmethod
    def from_dict(cls, data):
        """Create lookup table from dictionary"""
        lut = cls(data.get('name', 'Untitled'))
        lut.created_date = data.get('created_date', datetime.datetime.now().isoformat())
        lut.source_files = data.get('source_files', [])
        lut.metadata = data.get('metadata', {})
        lut.positions = data.get('raw_positions', [])
        lut.distances = data.get('raw_distances', [])
        
        if 'compiled_positions' in data:
            lut.compiled_positions = data['compiled_positions']
            lut.compiled_distances = data['compiled_distances']
            lut.compiled_std = data.get('compiled_std', [])
            lut.compiled_count = data.get('compiled_count', [])
        
        return lut
    
    def save(self, filepath):
        """Save lookup table to JSON file"""
        with open(filepath, 'w', encoding='utf-8') as f:
            json.dump(self.to_dict(), f, indent=2)
    
    @classmethod
    def load(cls, filepath):
        """Load lookup table from JSON file"""
        with open(filepath, 'r', encoding='utf-8') as f:
            data = json.load(f)
        return cls.from_dict(data)


class LookupTableGUI:
    """Main GUI application for lookup table management"""
    
    def __init__(self, root):
        self.root = root
        self.root.title("Sensor Lookup Table Tool")
        self.root.geometry("1200x800")
        
        # Data storage
        self.lookup_tables = {}  # name -> LookupTable
        self.current_lut = None
        self.loaded_files = []
        self.pending_data = {'positions': [], 'distances': [], 'files': []}
        
        # Default data directory
        self.data_dir = os.path.join(os.path.dirname(__file__), "data")
        
        self.setup_ui()
    
    def setup_ui(self):
        """Setup the main UI"""
        # Create main paned window
        self.main_paned = ttk.PanedWindow(self.root, orient=tk.HORIZONTAL)
        self.main_paned.pack(fill=tk.BOTH, expand=True, padx=5, pady=5)
        
        # Left panel - File browser and LUT list
        self.left_frame = ttk.Frame(self.main_paned, width=300)
        self.main_paned.add(self.left_frame, weight=1)
        
        # Right panel - Data view and plots
        self.right_frame = ttk.Frame(self.main_paned)
        self.main_paned.add(self.right_frame, weight=3)
        
        self.setup_left_panel()
        self.setup_right_panel()
        self.setup_menu()
    
    def setup_menu(self):
        """Setup menu bar"""
        menubar = tk.Menu(self.root)
        self.root.config(menu=menubar)
        
        # File menu
        file_menu = tk.Menu(menubar, tearoff=0)
        menubar.add_cascade(label="File", menu=file_menu)
        file_menu.add_command(label="Open Data Directory...", command=self.open_data_directory)
        file_menu.add_separator()
        file_menu.add_command(label="Load Lookup Table...", command=self.load_lut_file)
        file_menu.add_command(label="Save Lookup Table...", command=self.save_lut_file)
        file_menu.add_separator()
        file_menu.add_command(label="Export to C Header...", command=self.export_to_header)
        file_menu.add_separator()
        file_menu.add_command(label="Exit", command=self.root.quit)
        
        # Tools menu
        tools_menu = tk.Menu(menubar, tearoff=0)
        menubar.add_cascade(label="Tools", menu=tools_menu)
        tools_menu.add_command(label="Test Lookup Value...", command=self.test_lookup_dialog)
        tools_menu.add_command(label="Batch Correction...", command=self.batch_correction_dialog)
    
    def setup_left_panel(self):
        """Setup left panel with file browser and LUT list"""
        # Notebook for tabs
        notebook = ttk.Notebook(self.left_frame)
        notebook.pack(fill=tk.BOTH, expand=True)
        
        # Tab 1: Data Directory
        files_frame = ttk.Frame(notebook)
        notebook.add(files_frame, text="Data Directory")
        
        # Directory selection
        dir_frame = ttk.LabelFrame(files_frame, text="Data Directory")
        dir_frame.pack(fill=tk.X, padx=5, pady=5)
        
        self.dir_path_var = tk.StringVar(value=self.data_dir)
        dir_entry = ttk.Entry(dir_frame, textvariable=self.dir_path_var, state='readonly')
        dir_entry.pack(fill=tk.X, padx=5, pady=2)
        
        dir_btn_frame = ttk.Frame(dir_frame)
        dir_btn_frame.pack(fill=tk.X, padx=5, pady=5)
        ttk.Button(dir_btn_frame, text="Browse...", command=self.open_data_directory).pack(side=tk.LEFT)
        ttk.Button(dir_btn_frame, text="Refresh", command=self.refresh_file_list).pack(side=tk.LEFT, padx=5)
        
        # Load all button
        load_frame = ttk.LabelFrame(files_frame, text="Load Data")
        load_frame.pack(fill=tk.X, padx=5, pady=5)
        
        ttk.Button(load_frame, text="Load All Files in Directory", 
                   command=self.load_all_from_directory).pack(fill=tk.X, padx=5, pady=5)
        ttk.Button(load_frame, text="Clear Pending Data", 
                   command=self.clear_pending_data).pack(fill=tk.X, padx=5, pady=2)
        
        # File tree (shows what's available)
        tree_frame = ttk.LabelFrame(files_frame, text="Available Files")
        tree_frame.pack(fill=tk.BOTH, expand=True, padx=5, pady=5)
        
        self.file_tree = ttk.Treeview(tree_frame, selectmode='extended')
        self.file_tree.heading('#0', text='Excel Files')
        self.file_tree.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        
        scrollbar = ttk.Scrollbar(tree_frame, orient=tk.VERTICAL, command=self.file_tree.yview)
        scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
        self.file_tree.configure(yscrollcommand=scrollbar.set)
        
        # Tab 2: Lookup Tables
        lut_frame = ttk.Frame(notebook)
        notebook.add(lut_frame, text="Lookup Tables")
        
        # LUT list
        lut_list_frame = ttk.Frame(lut_frame)
        lut_list_frame.pack(fill=tk.BOTH, expand=True, padx=5, pady=5)
        
        self.lut_listbox = tk.Listbox(lut_list_frame, selectmode=tk.SINGLE)
        self.lut_listbox.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        self.lut_listbox.bind('<<ListboxSelect>>', self.on_lut_select)
        
        lut_scroll = ttk.Scrollbar(lut_list_frame, orient=tk.VERTICAL, command=self.lut_listbox.yview)
        lut_scroll.pack(side=tk.RIGHT, fill=tk.Y)
        self.lut_listbox.configure(yscrollcommand=lut_scroll.set)
        
        # LUT action buttons
        lut_btn_frame = ttk.Frame(lut_frame)
        lut_btn_frame.pack(fill=tk.X, padx=5, pady=5)
        
        ttk.Button(lut_btn_frame, text="New LUT", command=self.create_new_lut).pack(side=tk.LEFT)
        ttk.Button(lut_btn_frame, text="Delete", command=self.delete_lut).pack(side=tk.LEFT, padx=5)
        
        # Initialize file list
        self.refresh_file_list()
    
    def setup_right_panel(self):
        """Setup right panel with data view and plots"""
        # Notebook for different views
        self.right_notebook = ttk.Notebook(self.right_frame)
        self.right_notebook.pack(fill=tk.BOTH, expand=True)
        
        # Tab 1: Pending Data
        self.pending_frame = ttk.Frame(self.right_notebook)
        self.right_notebook.add(self.pending_frame, text="Pending Data")
        self.setup_pending_tab()
        
        # Tab 2: Compiled LUT
        self.compiled_frame = ttk.Frame(self.right_notebook)
        self.right_notebook.add(self.compiled_frame, text="Compiled LUT")
        self.setup_compiled_tab()
        
        # Tab 3: Test/Apply
        self.test_frame = ttk.Frame(self.right_notebook)
        self.right_notebook.add(self.test_frame, text="Test & Apply")
        self.setup_test_tab()
        
        # Tab 4: Auto-Optimize
        self.optimize_frame = ttk.Frame(self.right_notebook)
        self.right_notebook.add(self.optimize_frame, text="Auto-Optimize")
        self.setup_optimize_tab()
    
    def setup_pending_tab(self):
        """Setup the pending data tab"""
        # Info frame
        info_frame = ttk.LabelFrame(self.pending_frame, text="Pending Data Info")
        info_frame.pack(fill=tk.X, padx=5, pady=5)
        
        self.pending_info_label = ttk.Label(info_frame, text="No data loaded")
        self.pending_info_label.pack(padx=5, pady=5)
        
        # Loaded files list
        files_frame = ttk.LabelFrame(self.pending_frame, text="Loaded Files")
        files_frame.pack(fill=tk.X, padx=5, pady=5)
        
        self.loaded_files_text = tk.Text(files_frame, height=4, state=tk.DISABLED)
        self.loaded_files_text.pack(fill=tk.X, padx=5, pady=5)
        
        # Compile options
        options_frame = ttk.LabelFrame(self.pending_frame, text="Compile Options")
        options_frame.pack(fill=tk.X, padx=5, pady=5)
        
        # Bin size
        bin_frame = ttk.Frame(options_frame)
        bin_frame.pack(fill=tk.X, padx=5, pady=2)
        ttk.Label(bin_frame, text="Bin Size (mm):").pack(side=tk.LEFT)
        self.bin_size_var = tk.StringVar(value="1.0")
        ttk.Entry(bin_frame, textvariable=self.bin_size_var, width=10).pack(side=tk.LEFT, padx=5)
        
        # Method
        method_frame = ttk.Frame(options_frame)
        method_frame.pack(fill=tk.X, padx=5, pady=2)
        ttk.Label(method_frame, text="Method:").pack(side=tk.LEFT)
        self.method_var = tk.StringVar(value="average")
        ttk.Combobox(method_frame, textvariable=self.method_var, 
                     values=["average", "median"], state="readonly", width=15).pack(side=tk.LEFT, padx=5)
        
        # LUT name
        name_frame = ttk.Frame(options_frame)
        name_frame.pack(fill=tk.X, padx=5, pady=2)
        ttk.Label(name_frame, text="LUT Name:").pack(side=tk.LEFT)
        self.lut_name_var = tk.StringVar(value="New_LUT")
        ttk.Entry(name_frame, textvariable=self.lut_name_var, width=20).pack(side=tk.LEFT, padx=5)
        
        # Smoothing options
        smooth_label_frame = ttk.LabelFrame(options_frame, text="Smoothing (Optional)")
        smooth_label_frame.pack(fill=tk.X, padx=5, pady=5)
        
        # Enable smoothing checkbox
        self.smooth_enabled_var = tk.BooleanVar(value=False)
        ttk.Checkbutton(smooth_label_frame, text="Enable Smoothing", 
                       variable=self.smooth_enabled_var).pack(anchor=tk.W, padx=5, pady=2)
        
        # Smoothing method
        smooth_method_frame = ttk.Frame(smooth_label_frame)
        smooth_method_frame.pack(fill=tk.X, padx=5, pady=2)
        ttk.Label(smooth_method_frame, text="Method:").pack(side=tk.LEFT)
        self.smooth_method_var = tk.StringVar(value="moving_average")
        ttk.Combobox(smooth_method_frame, textvariable=self.smooth_method_var, 
                     values=["moving_average", "savgol", "gaussian"], 
                     state="readonly", width=15).pack(side=tk.LEFT, padx=5)
        
        # Smoothing window
        smooth_window_frame = ttk.Frame(smooth_label_frame)
        smooth_window_frame.pack(fill=tk.X, padx=5, pady=2)
        ttk.Label(smooth_window_frame, text="Window Size:").pack(side=tk.LEFT)
        self.smooth_window_var = tk.StringVar(value="5")
        ttk.Entry(smooth_window_frame, textvariable=self.smooth_window_var, width=10).pack(side=tk.LEFT, padx=5)
        ttk.Label(smooth_window_frame, text="(larger = smoother)").pack(side=tk.LEFT)
        
        # Compile button
        ttk.Button(options_frame, text="Compile Lookup Table", 
                   command=self.compile_pending_data).pack(pady=10)
        
        # Plot frame
        plot_frame = ttk.LabelFrame(self.pending_frame, text="Raw Data Preview")
        plot_frame.pack(fill=tk.BOTH, expand=True, padx=5, pady=5)
        
        self.pending_fig, self.pending_ax = plt.subplots(figsize=(8, 4))
        self.pending_canvas = FigureCanvasTkAgg(self.pending_fig, plot_frame)
        self.pending_canvas.get_tk_widget().pack(fill=tk.BOTH, expand=True)
        
        toolbar_frame = ttk.Frame(plot_frame)
        toolbar_frame.pack(fill=tk.X)
        NavigationToolbar2Tk(self.pending_canvas, toolbar_frame)
    
    def setup_compiled_tab(self):
        """Setup the compiled LUT tab"""
        # Info frame
        info_frame = ttk.LabelFrame(self.compiled_frame, text="Lookup Table Info")
        info_frame.pack(fill=tk.X, padx=5, pady=5)
        
        self.compiled_info_label = ttk.Label(info_frame, text="No lookup table selected")
        self.compiled_info_label.pack(padx=5, pady=5)
        
        # Data table
        table_frame = ttk.LabelFrame(self.compiled_frame, text="LUT Data")
        table_frame.pack(fill=tk.BOTH, expand=True, padx=5, pady=5)
        
        # Treeview for data
        columns = ('position', 'distance', 'std', 'count')
        self.lut_tree = ttk.Treeview(table_frame, columns=columns, show='headings', height=10)
        
        self.lut_tree.heading('position', text='Position (mm)')
        self.lut_tree.heading('distance', text='Distance (mm)')
        self.lut_tree.heading('std', text='Std Dev')
        self.lut_tree.heading('count', text='Samples')
        
        self.lut_tree.column('position', width=100)
        self.lut_tree.column('distance', width=100)
        self.lut_tree.column('std', width=80)
        self.lut_tree.column('count', width=80)
        
        self.lut_tree.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        
        tree_scroll = ttk.Scrollbar(table_frame, orient=tk.VERTICAL, command=self.lut_tree.yview)
        tree_scroll.pack(side=tk.RIGHT, fill=tk.Y)
        self.lut_tree.configure(yscrollcommand=tree_scroll.set)
        
        # Plot frame
        plot_frame = ttk.LabelFrame(self.compiled_frame, text="LUT Visualization")
        plot_frame.pack(fill=tk.BOTH, expand=True, padx=5, pady=5)
        
        self.compiled_fig, self.compiled_axes = plt.subplots(1, 2, figsize=(10, 4))
        self.compiled_canvas = FigureCanvasTkAgg(self.compiled_fig, plot_frame)
        self.compiled_canvas.get_tk_widget().pack(fill=tk.BOTH, expand=True)
        
        toolbar_frame = ttk.Frame(plot_frame)
        toolbar_frame.pack(fill=tk.X)
        NavigationToolbar2Tk(self.compiled_canvas, toolbar_frame)
    
    def setup_test_tab(self):
        """Setup the test & apply tab"""
        # Single value test
        test_frame = ttk.LabelFrame(self.test_frame, text="Test Lookup (Sensor Distance → True Position)")
        test_frame.pack(fill=tk.X, padx=5, pady=5)
        
        input_frame = ttk.Frame(test_frame)
        input_frame.pack(fill=tk.X, padx=5, pady=5)
        
        ttk.Label(input_frame, text="Sensor Distance (mm):").pack(side=tk.LEFT)
        self.test_input_var = tk.StringVar()
        ttk.Entry(input_frame, textvariable=self.test_input_var, width=15).pack(side=tk.LEFT, padx=5)
        ttk.Button(input_frame, text="Get True Position", command=self.test_single_lookup).pack(side=tk.LEFT)
        
        self.test_result_label = ttk.Label(test_frame, text="Result: -")
        self.test_result_label.pack(padx=5, pady=5)
        
        # Lookup Table Correction Application
        correction_frame = ttk.LabelFrame(self.test_frame, text="Apply Lookup Table Correction to TDS Data")
        correction_frame.pack(fill=tk.X, padx=5, pady=5)
        
        # Load TDS file for correction
        load_frame = ttk.Frame(correction_frame)
        load_frame.pack(fill=tk.X, padx=5, pady=5)
        
        ttk.Button(load_frame, text="Load TDS File...", command=self.load_tds_for_correction).pack(side=tk.LEFT)
        self.tds_file_label = ttk.Label(load_frame, text="No file loaded")
        self.tds_file_label.pack(side=tk.LEFT, padx=10)
        
        # LUT selection display
        lut_frame = ttk.Frame(correction_frame)
        lut_frame.pack(fill=tk.X, padx=5, pady=5)
        
        ttk.Label(lut_frame, text="Active Lookup Table:").pack(side=tk.LEFT)
        self.active_lut_label = ttk.Label(lut_frame, text="None selected", foreground="red")
        self.active_lut_label.pack(side=tk.LEFT, padx=5)
        
        # Correction method selection
        method_frame = ttk.LabelFrame(correction_frame, text="Correction Method")
        method_frame.pack(fill=tk.X, padx=5, pady=5)
        
        self.correction_method_var = tk.StringVar(value="reverse_lookup")
        
        # Create scrollable frame for methods
        method_canvas = tk.Canvas(method_frame, height=180)
        method_scrollbar = ttk.Scrollbar(method_frame, orient="vertical", command=method_canvas.yview)
        method_inner_frame = ttk.Frame(method_canvas)
        
        method_inner_frame.bind(
            "<Configure>",
            lambda e: method_canvas.configure(scrollregion=method_canvas.bbox("all"))
        )
        
        method_canvas.create_window((0, 0), window=method_inner_frame, anchor="nw")
        method_canvas.configure(yscrollcommand=method_scrollbar.set)
        
        # Linear methods
        ttk.Label(method_inner_frame, text="Linear Methods:", font=('TkDefaultFont', 9, 'bold')).pack(anchor=tk.W, padx=5, pady=(2,0))
        
        ttk.Radiobutton(method_inner_frame, text="Reverse Lookup (Linear)", 
                       variable=self.correction_method_var, 
                       value="reverse_lookup").pack(anchor=tk.W, padx=5, pady=1)
        ttk.Label(method_inner_frame, text="   Fast, linear interpolation between LUT points",
                 font=('TkDefaultFont', 8), foreground='gray').pack(anchor=tk.W, padx=20)
        
        ttk.Radiobutton(method_inner_frame, text="Error Subtraction (Linear)", 
                       variable=self.correction_method_var, 
                       value="error_subtraction").pack(anchor=tk.W, padx=5, pady=1)
        ttk.Label(method_inner_frame, text="   Interpolates error curve, linear between points",
                 font=('TkDefaultFont', 8), foreground='gray').pack(anchor=tk.W, padx=20)
        
        # Smooth methods
        ttk.Label(method_inner_frame, text="\nSmooth Methods:", font=('TkDefaultFont', 9, 'bold')).pack(anchor=tk.W, padx=5, pady=(5,0))
        
        ttk.Radiobutton(method_inner_frame, text="Cubic Spline", 
                       variable=self.correction_method_var, 
                       value="spline").pack(anchor=tk.W, padx=5, pady=1)
        ttk.Label(method_inner_frame, text="   Smooth curves, continuous derivatives (requires scipy)",
                 font=('TkDefaultFont', 8), foreground='gray').pack(anchor=tk.W, padx=20)
        
        ttk.Radiobutton(method_inner_frame, text="Polynomial Fit (degree 3)", 
                       variable=self.correction_method_var, 
                       value="polynomial").pack(anchor=tk.W, padx=5, pady=1)
        ttk.Label(method_inner_frame, text="   Fits polynomial to error curve, very smooth",
                 font=('TkDefaultFont', 8), foreground='gray').pack(anchor=tk.W, padx=20)
        
        ttk.Radiobutton(method_inner_frame, text="Weighted Average (k=3)", 
                       variable=self.correction_method_var, 
                       value="weighted").pack(anchor=tk.W, padx=5, pady=1)
        ttk.Label(method_inner_frame, text="   Uses 3 nearest points with distance weighting",
                 font=('TkDefaultFont', 8), foreground='gray').pack(anchor=tk.W, padx=20)
        
        # Simple methods
        ttk.Label(method_inner_frame, text="\nSimple Methods:", font=('TkDefaultFont', 9, 'bold')).pack(anchor=tk.W, padx=5, pady=(5,0))
        
        ttk.Radiobutton(method_inner_frame, text="Nearest Neighbor", 
                       variable=self.correction_method_var, 
                       value="nearest").pack(anchor=tk.W, padx=5, pady=1)
        ttk.Label(method_inner_frame, text="   Uses closest LUT point, no interpolation",
                 font=('TkDefaultFont', 8), foreground='gray').pack(anchor=tk.W, padx=20)
        
        method_canvas.pack(side=tk.LEFT, fill=tk.BOTH, expand=True, padx=5, pady=5)
        method_scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
        
        ttk.Button(method_frame, text="Compare All Methods...", 
                  command=self.compare_correction_methods).pack(pady=5)
        
        # Method Parameters
        params_frame = ttk.LabelFrame(correction_frame, text="Method Parameters")
        params_frame.pack(fill=tk.X, padx=5, pady=5)
        
        # Polynomial degree
        poly_frame = ttk.Frame(params_frame)
        poly_frame.pack(fill=tk.X, padx=5, pady=2)
        ttk.Label(poly_frame, text="Polynomial Degree:").pack(side=tk.LEFT)
        self.poly_degree_var = tk.StringVar(value="3")
        poly_spinbox = ttk.Spinbox(poly_frame, from_=1, to=10, textvariable=self.poly_degree_var, width=8)
        poly_spinbox.pack(side=tk.LEFT, padx=5)
        ttk.Label(poly_frame, text="(for Polynomial method)", 
                 font=('TkDefaultFont', 8), foreground='gray').pack(side=tk.LEFT)
        
        # Weighted k neighbors
        k_frame = ttk.Frame(params_frame)
        k_frame.pack(fill=tk.X, padx=5, pady=2)
        ttk.Label(k_frame, text="K Neighbors:").pack(side=tk.LEFT)
        self.k_neighbors_var = tk.StringVar(value="3")
        k_spinbox = ttk.Spinbox(k_frame, from_=1, to=20, textvariable=self.k_neighbors_var, width=8)
        k_spinbox.pack(side=tk.LEFT, padx=5)
        ttk.Label(k_frame, text="(for Weighted Average method)", 
                 font=('TkDefaultFont', 8), foreground='gray').pack(side=tk.LEFT)
        
        # Weighted power
        power_frame = ttk.Frame(params_frame)
        power_frame.pack(fill=tk.X, padx=5, pady=2)
        ttk.Label(power_frame, text="Distance Power:").pack(side=tk.LEFT)
        self.distance_power_var = tk.StringVar(value="1.0")
        power_entry = ttk.Entry(power_frame, textvariable=self.distance_power_var, width=8)
        power_entry.pack(side=tk.LEFT, padx=5)
        ttk.Label(power_frame, text="(1.0=linear, 2.0=quadratic, for Weighted)", 
                 font=('TkDefaultFont', 8), foreground='gray').pack(side=tk.LEFT)
        
        # Apply correction button
        apply_frame = ttk.Frame(correction_frame)
        apply_frame.pack(fill=tk.X, padx=5, pady=5)
        
        ttk.Button(apply_frame, text="Apply LUT Correction", command=self.apply_lut_correction).pack(side=tk.LEFT)
        ttk.Button(apply_frame, text="Save Corrected TDS...", command=self.save_corrected_tds).pack(side=tk.LEFT, padx=5)
        ttk.Button(apply_frame, text="View Comparison Plot", command=self.show_correction_comparison).pack(side=tk.LEFT, padx=5)
        
        # Stats display
        self.correction_stats_label = ttk.Label(correction_frame, text="", justify=tk.LEFT)
        self.correction_stats_label.pack(padx=5, pady=5)
        
        # Batch correction
        batch_frame = ttk.LabelFrame(self.test_frame, text="Batch Correction")
        batch_frame.pack(fill=tk.BOTH, expand=True, padx=5, pady=5)
        
        btn_frame = ttk.Frame(batch_frame)
        btn_frame.pack(fill=tk.X, padx=5, pady=5)
        
        ttk.Button(btn_frame, text="Load CSV Data...", command=self.load_batch_data).pack(side=tk.LEFT)
        ttk.Button(btn_frame, text="Apply Correction", command=self.apply_batch_correction).pack(side=tk.LEFT, padx=5)
        ttk.Button(btn_frame, text="Save Corrected Data...", command=self.save_corrected_data).pack(side=tk.LEFT)
        
        # Batch data preview
        self.batch_text = tk.Text(batch_frame, height=10)
        self.batch_text.pack(fill=tk.BOTH, expand=True, padx=5, pady=5)
        
        self.batch_data = None
        self.corrected_data = None
        self.tds_data = None
        self.corrected_tds_data = None
        self.optimization_results = None
    
    def setup_optimize_tab(self):
        """Setup the auto-optimization tab"""
        # Info frame
        info_frame = ttk.LabelFrame(self.optimize_frame, text="Auto-Optimization")
        info_frame.pack(fill=tk.X, padx=5, pady=5)
        
        info_text = (
            "This tool tests all correction methods with various parameter combinations\n"
            "to find the best possible improvement for your data.\n\n"
            "You must have TDS data loaded in the 'Test & Apply' tab first."
        )
        ttk.Label(info_frame, text=info_text, justify=tk.LEFT).pack(padx=5, pady=5)
        
        # Configuration frame
        config_frame = ttk.LabelFrame(self.optimize_frame, text="Search Configuration")
        config_frame.pack(fill=tk.X, padx=5, pady=5)
        
        # Polynomial degree range
        poly_range_frame = ttk.Frame(config_frame)
        poly_range_frame.pack(fill=tk.X, padx=5, pady=2)
        ttk.Label(poly_range_frame, text="Polynomial Degrees:").pack(side=tk.LEFT)
        self.opt_poly_min_var = tk.StringVar(value="1")
        ttk.Spinbox(poly_range_frame, from_=1, to=10, textvariable=self.opt_poly_min_var, width=5).pack(side=tk.LEFT, padx=2)
        ttk.Label(poly_range_frame, text="to").pack(side=tk.LEFT)
        self.opt_poly_max_var = tk.StringVar(value="5")
        ttk.Spinbox(poly_range_frame, from_=1, to=10, textvariable=self.opt_poly_max_var, width=5).pack(side=tk.LEFT, padx=2)
        
        # K neighbors range
        k_range_frame = ttk.Frame(config_frame)
        k_range_frame.pack(fill=tk.X, padx=5, pady=2)
        ttk.Label(k_range_frame, text="K Neighbors:").pack(side=tk.LEFT)
        self.opt_k_min_var = tk.StringVar(value="1")
        ttk.Spinbox(k_range_frame, from_=1, to=20, textvariable=self.opt_k_min_var, width=5).pack(side=tk.LEFT, padx=2)
        ttk.Label(k_range_frame, text="to").pack(side=tk.LEFT)
        self.opt_k_max_var = tk.StringVar(value="7")
        ttk.Spinbox(k_range_frame, from_=1, to=20, textvariable=self.opt_k_max_var, width=5).pack(side=tk.LEFT, padx=2)
        
        # Distance power range
        power_range_frame = ttk.Frame(config_frame)
        power_range_frame.pack(fill=tk.X, padx=5, pady=2)
        ttk.Label(power_range_frame, text="Distance Powers:").pack(side=tk.LEFT)
        self.opt_power_values_var = tk.StringVar(value="0.5, 1.0, 1.5, 2.0")
        ttk.Entry(power_range_frame, textvariable=self.opt_power_values_var, width=30).pack(side=tk.LEFT, padx=5)
        ttk.Label(power_range_frame, text="(comma-separated)", font=('TkDefaultFont', 8), foreground='gray').pack(side=tk.LEFT)
        
        # Run button
        btn_frame = ttk.Frame(config_frame)
        btn_frame.pack(fill=tk.X, padx=5, pady=10)
        ttk.Button(btn_frame, text="Run Optimization", command=self.run_optimization).pack(side=tk.LEFT, padx=5)
        self.opt_progress_label = ttk.Label(btn_frame, text="")
        self.opt_progress_label.pack(side=tk.LEFT, padx=10)
        
        # Results frame
        results_frame = ttk.LabelFrame(self.optimize_frame, text="Optimization Results")
        results_frame.pack(fill=tk.BOTH, expand=True, padx=5, pady=5)
        
        # Results table
        table_frame = ttk.Frame(results_frame)
        table_frame.pack(fill=tk.BOTH, expand=True, padx=5, pady=5)
        
        columns = ('rank', 'method', 'parameters', 'mean_error', 'improvement', 'std_error')
        self.opt_tree = ttk.Treeview(table_frame, columns=columns, show='headings', height=15)
        
        self.opt_tree.heading('rank', text='Rank')
        self.opt_tree.heading('method', text='Method')
        self.opt_tree.heading('parameters', text='Parameters')
        self.opt_tree.heading('mean_error', text='Mean Error (mm)')
        self.opt_tree.heading('improvement', text='Improvement (%)')
        self.opt_tree.heading('std_error', text='Std Dev (mm)')
        
        self.opt_tree.column('rank', width=50)
        self.opt_tree.column('method', width=150)
        self.opt_tree.column('parameters', width=200)
        self.opt_tree.column('mean_error', width=120)
        self.opt_tree.column('improvement', width=120)
        self.opt_tree.column('std_error', width=100)
        
        self.opt_tree.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        
        tree_scroll = ttk.Scrollbar(table_frame, orient=tk.VERTICAL, command=self.opt_tree.yview)
        tree_scroll.pack(side=tk.RIGHT, fill=tk.Y)
        self.opt_tree.configure(yscrollcommand=tree_scroll.set)
        
        # Action buttons
        action_frame = ttk.Frame(results_frame)
        action_frame.pack(fill=tk.X, padx=5, pady=5)
        
        ttk.Button(action_frame, text="Apply Best Method", command=self.apply_best_method).pack(side=tk.LEFT, padx=5)
        ttk.Button(action_frame, text="Visualize Top 5", command=self.visualize_top_methods).pack(side=tk.LEFT, padx=5)
        ttk.Button(action_frame, text="Export Results...", command=self.export_optimization_results).pack(side=tk.LEFT, padx=5)
    
    def open_data_directory(self):
        """Open a data directory"""
        directory = filedialog.askdirectory(initialdir=self.data_dir, title="Select Data Directory")
        if directory:
            self.data_dir = directory
            self.dir_path_var.set(directory)
            self.refresh_file_list()
    
    def refresh_file_list(self):
        """Refresh the file tree with available data files"""
        # Clear existing items
        for item in self.file_tree.get_children():
            self.file_tree.delete(item)
        
        if not os.path.exists(self.data_dir):
            return
        
        # Walk through data directory
        for root, dirs, files in os.walk(self.data_dir):
            rel_path = os.path.relpath(root, self.data_dir)
            
            if rel_path == '.':
                parent = ''
            else:
                parent = self.file_tree.insert('', 'end', rel_path, text=rel_path, open=True)
            
            for f in sorted(files):
                if f.endswith('.xlsx') and f.startswith('TDS_'):
                    file_path = os.path.join(root, f)
                    item_id = file_path
                    if parent:
                        self.file_tree.insert(parent, 'end', item_id, text=f)
                    else:
                        self.file_tree.insert('', 'end', item_id, text=f)
    
    def load_all_from_directory(self):
        """Load all Excel files from the data directory"""
        if not os.path.exists(self.data_dir):
            messagebox.showwarning("Warning", "Data directory does not exist")
            return
        
        # Clear existing pending data
        self.clear_pending_data()
        
        file_count = 0
        # Walk through data directory and load all TDS_*.xlsx files
        for root, dirs, files in os.walk(self.data_dir):
            for f in sorted(files):
                if f.endswith('.xlsx') and f.startswith('TDS_'):
                    filepath = os.path.join(root, f)
                    self.load_xlsx_file(filepath)
                    file_count += 1
        
        if file_count == 0:
            messagebox.showinfo("Info", "No TDS_*.xlsx files found in directory")
        else:
            messagebox.showinfo("Success", f"Loaded {file_count} files with {len(self.pending_data['positions'])} data points")
        
        self.update_pending_display()
    
    def load_xlsx_file(self, filepath):
        """Load an Excel file and add to pending data"""
        try:
            wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
            
            # Try to find RAW_DATA sheet first, otherwise use active sheet
            if 'RAW_DATA' in wb.sheetnames:
                ws = wb['RAW_DATA']
            else:
                ws = wb.active
            
            row_count = 0
            for row in ws.iter_rows(min_row=1, values_only=True):
                if row and len(row) >= 4:
                    try:
                        # Excel format: distance, temp, position, delta, ...
                        distance = float(row[0]) if row[0] is not None else None
                        position = float(row[2]) if row[2] is not None else None
                        
                        if distance is not None and position is not None:
                            self.pending_data['positions'].append(position)
                            self.pending_data['distances'].append(distance)
                            row_count += 1
                    except (ValueError, TypeError):
                        continue
            
            wb.close()
            
            if filepath not in self.pending_data['files']:
                self.pending_data['files'].append(filepath)
            
            print(f"Loaded: {filepath} ({row_count} data points)")
            
        except Exception as e:
            messagebox.showerror("Error", f"Failed to load {filepath}: {str(e)}")
    
    def clear_pending_data(self):
        """Clear all pending data"""
        self.pending_data = {'positions': [], 'distances': [], 'files': []}
        self.update_pending_display()
    
    def update_pending_display(self):
        """Update the pending data display"""
        count = len(self.pending_data['positions'])
        file_count = len(self.pending_data['files'])
        
        self.pending_info_label.config(
            text=f"Data Points: {count} | Files Loaded: {file_count}"
        )
        
        # Update loaded files text
        self.loaded_files_text.config(state=tk.NORMAL)
        self.loaded_files_text.delete(1.0, tk.END)
        for f in self.pending_data['files']:
            self.loaded_files_text.insert(tk.END, os.path.basename(f) + "\n")
        self.loaded_files_text.config(state=tk.DISABLED)
        
        # Update plot
        self.pending_ax.clear()
        if count > 0:
            self.pending_ax.scatter(self.pending_data['positions'], 
                                    self.pending_data['distances'],
                                    alpha=0.5, s=10)
            self.pending_ax.set_xlabel('Position (mm)')
            self.pending_ax.set_ylabel('Sensor Distance (mm)')
            self.pending_ax.set_title('Raw Data (Position vs Distance)')
            self.pending_ax.grid(True, alpha=0.3)
            
            # Add ideal line
            min_val = min(min(self.pending_data['positions']), min(self.pending_data['distances']))
            max_val = max(max(self.pending_data['positions']), max(self.pending_data['distances']))
            self.pending_ax.plot([min_val, max_val], [min_val, max_val], 'r--', label='Ideal (y=x)')
            self.pending_ax.legend()
        
        self.pending_fig.tight_layout()
        self.pending_canvas.draw()
    
    def compile_pending_data(self):
        """Compile pending data into a lookup table"""
        if not self.pending_data['positions']:
            messagebox.showwarning("Warning", "No data to compile")
            return
        
        try:
            bin_size = float(self.bin_size_var.get())
        except ValueError:
            messagebox.showerror("Error", "Invalid bin size")
            return
        
        name = self.lut_name_var.get().strip()
        if not name:
            name = f"LUT_{datetime.datetime.now().strftime('%Y%m%d_%H%M%S')}"
        
        # Create new lookup table
        lut = LookupTable(name)
        lut.add_data(self.pending_data['positions'], self.pending_data['distances'])
        lut.source_files = self.pending_data['files'].copy()
        
        # Get smoothing parameters
        smooth_enabled = self.smooth_enabled_var.get()
        smooth_method = self.smooth_method_var.get()
        try:
            smooth_window = int(self.smooth_window_var.get())
            if smooth_window < 2:
                smooth_window = 2
        except ValueError:
            smooth_window = 5
        
        # Compile
        method = self.method_var.get()
        if lut.compile(bin_size=bin_size, method=method, smooth=smooth_enabled, 
                      smooth_method=smooth_method, smooth_window=smooth_window):
            self.lookup_tables[name] = lut
            self.current_lut = lut
            self.update_lut_list()
            self.update_compiled_display()
            
            # Auto-save lookup table to data directory
            try:
                lut_dir = os.path.join(self.data_dir, "lookup_tables")
                os.makedirs(lut_dir, exist_ok=True)
                
                # Save JSON
                json_path = os.path.join(lut_dir, f"{name}.json")
                lut.save(json_path)
                
                # Save Python module for easy import
                py_path = os.path.join(lut_dir, f"{name}.py")
                self.write_python_module(py_path, lut)
                
                # Save C header
                h_path = os.path.join(lut_dir, f"{name}.h")
                self.write_c_header(h_path)
                
                messagebox.showinfo("Success", 
                    f"Lookup table '{name}' created with {len(lut.compiled_positions)} entries\n\n"
                    f"Saved to: {lut_dir}\n"
                    f"- {name}.json (data file)\n"
                    f"- {name}.py (Python module)\n"
                    f"- {name}.h (C header)")
            except Exception as e:
                messagebox.showwarning("Warning", 
                    f"Lookup table created but failed to auto-save: {str(e)}")
            
            # Switch to compiled tab
            self.right_notebook.select(1)
        else:
            messagebox.showerror("Error", "Failed to compile lookup table")
    
    def update_lut_list(self):
        """Update the lookup table listbox"""
        self.lut_listbox.delete(0, tk.END)
        for name in sorted(self.lookup_tables.keys()):
            self.lut_listbox.insert(tk.END, name)
    
    def on_lut_select(self, event):
        """Handle LUT selection"""
        selection = self.lut_listbox.curselection()
        if selection:
            name = self.lut_listbox.get(selection[0])
            self.current_lut = self.lookup_tables.get(name)
            self.update_compiled_display()
            self.update_active_lut_label()
    
    def update_compiled_display(self):
        """Update the compiled LUT display"""
        # Clear tree
        for item in self.lut_tree.get_children():
            self.lut_tree.delete(item)
        
        if not self.current_lut or not hasattr(self.current_lut, 'compiled_positions'):
            self.compiled_info_label.config(text="No lookup table selected")
            return
        
        lut = self.current_lut
        
        # Update info
        smooth_info = ""
        if lut.metadata.get('smoothed', False):
            smooth_info = f" | Smoothed: {lut.metadata.get('smooth_method', 'N/A')} (w={lut.metadata.get('smooth_window', 'N/A')})"
        
        info_text = (f"Name: {lut.name} | "
                    f"Entries: {len(lut.compiled_positions)} | "
                    f"Bin Size: {lut.metadata.get('bin_size', 'N/A')} mm | "
                    f"Method: {lut.metadata.get('method', 'N/A')}{smooth_info}")
        self.compiled_info_label.config(text=info_text)
        
        # Populate tree
        for i in range(len(lut.compiled_positions)):
            self.lut_tree.insert('', 'end', values=(
                f"{lut.compiled_positions[i]:.2f}",
                f"{lut.compiled_distances[i]:.2f}",
                f"{lut.compiled_std[i]:.3f}" if lut.compiled_std else "N/A",
                lut.compiled_count[i] if lut.compiled_count else "N/A"
            ))
        
        # Update plots
        for ax in self.compiled_axes:
            ax.clear()
        
        # Plot 1: Position vs Distance with error bars
        ax1 = self.compiled_axes[0]
        if lut.compiled_std and any(s > 0 for s in lut.compiled_std):
            ax1.errorbar(lut.compiled_positions, lut.compiled_distances, 
                        yerr=lut.compiled_std, fmt='o-', capsize=3, markersize=4)
        else:
            ax1.plot(lut.compiled_positions, lut.compiled_distances, 'o-', markersize=4)
        
        # Add ideal line
        min_val = min(min(lut.compiled_positions), min(lut.compiled_distances))
        max_val = max(max(lut.compiled_positions), max(lut.compiled_distances))
        ax1.plot([min_val, max_val], [min_val, max_val], 'r--', alpha=0.7, label='Ideal (y=x)')
        
        ax1.set_xlabel('Position (mm)')
        ax1.set_ylabel('Sensor Distance (mm)')
        ax1.set_title('Lookup Table')
        ax1.grid(True, alpha=0.3)
        ax1.legend()
        
        # Plot 2: Error/Delta
        ax2 = self.compiled_axes[1]
        deltas = [lut.compiled_distances[i] - lut.compiled_positions[i] 
                  for i in range(len(lut.compiled_positions))]
        ax2.plot(lut.compiled_positions, deltas, 'o-', color='orange', markersize=4)
        ax2.axhline(y=0, color='r', linestyle='--', alpha=0.7)
        ax2.set_xlabel('Position (mm)')
        ax2.set_ylabel('Error (mm)')
        ax2.set_title('Sensor Error vs Position')
        ax2.grid(True, alpha=0.3)
        
        self.compiled_fig.tight_layout()
        self.compiled_canvas.draw()
    
    def create_new_lut(self):
        """Create a new empty lookup table"""
        name = f"LUT_{datetime.datetime.now().strftime('%Y%m%d_%H%M%S')}"
        self.lut_name_var.set(name)
        self.right_notebook.select(0)  # Switch to pending data tab
    
    def delete_lut(self):
        """Delete selected lookup table"""
        selection = self.lut_listbox.curselection()
        if selection:
            name = self.lut_listbox.get(selection[0])
            if messagebox.askyesno("Confirm", f"Delete lookup table '{name}'?"):
                del self.lookup_tables[name]
                if self.current_lut and self.current_lut.name == name:
                    self.current_lut = None
                self.update_lut_list()
                self.update_compiled_display()
    
    def load_lut_file(self):
        """Load a lookup table from file"""
        filepath = filedialog.askopenfilename(
            title="Load Lookup Table",
            filetypes=[("JSON files", "*.json"), ("All files", "*.*")]
        )
        if filepath:
            try:
                lut = LookupTable.load(filepath)
                self.lookup_tables[lut.name] = lut
                self.current_lut = lut
                self.update_lut_list()
                self.update_compiled_display()
                messagebox.showinfo("Success", f"Loaded lookup table: {lut.name}")
            except Exception as e:
                messagebox.showerror("Error", f"Failed to load: {str(e)}")
    
    def save_lut_file(self):
        """Save current lookup table to file"""
        if not self.current_lut:
            messagebox.showwarning("Warning", "No lookup table selected")
            return
        
        filepath = filedialog.asksaveasfilename(
            title="Save Lookup Table",
            defaultextension=".json",
            initialfile=f"{self.current_lut.name}.json",
            filetypes=[("JSON files", "*.json"), ("All files", "*.*")]
        )
        if filepath:
            try:
                self.current_lut.save(filepath)
                messagebox.showinfo("Success", f"Saved: {filepath}")
            except Exception as e:
                messagebox.showerror("Error", f"Failed to save: {str(e)}")
    
    def export_to_header(self):
        """Export current lookup table to C header file"""
        if not self.current_lut or not hasattr(self.current_lut, 'compiled_positions'):
            messagebox.showwarning("Warning", "No compiled lookup table selected")
            return
        
        filepath = filedialog.asksaveasfilename(
            title="Export to C Header",
            defaultextension=".h",
            initialfile=f"{self.current_lut.name}.h",
            filetypes=[("C Header files", "*.h"), ("All files", "*.*")]
        )
        if filepath:
            try:
                self.write_c_header(filepath)
                messagebox.showinfo("Success", f"Exported: {filepath}")
            except Exception as e:
                messagebox.showerror("Error", f"Failed to export: {str(e)}")
    
    def write_python_module(self, filepath, lut=None):
        """Write lookup table as importable Python module"""
        if lut is None:
            lut = self.current_lut
        
        with open(filepath, 'w', encoding='utf-8') as f:
            f.write('"""\n')
            f.write(f"Sensor Distance Lookup Table: {lut.name}\n")
            f.write(f"Generated: {datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n")
            f.write(f"Bin Size: {lut.metadata.get('bin_size', 'N/A')} mm\n")
            f.write(f"Method: {lut.metadata.get('method', 'N/A')}\n")
            f.write('\n')
            f.write('Usage:\n')
            f.write('    from {} import get_true_position\n'.format(os.path.splitext(os.path.basename(filepath))[0]))
            f.write('    \n')
            f.write('    sensor_distance = 150.5  # mm from sensor\n')
            f.write('    true_position = get_true_position(sensor_distance)\n')
            f.write('    print(f"True position: {true_position:.2f}mm")\n')
            f.write('"""\n\n')
            
            f.write('# Lookup table data\n')
            f.write(f"LUT_POSITIONS = {lut.compiled_positions}\n\n")
            f.write(f"LUT_DISTANCES = {lut.compiled_distances}\n\n")
            
            f.write('def get_true_position(sensor_distance):\n')
            f.write('    """\n')
            f.write('    Get the true position for a given sensor distance reading.\n')
            f.write('    \n')
            f.write('    Args:\n')
            f.write('        sensor_distance: Distance measured by the sensor (mm)\n')
            f.write('    \n')
            f.write('    Returns:\n')
            f.write('        True position (mm) or None if out of range\n')
            f.write('    """\n')
            f.write('    positions = LUT_POSITIONS\n')
            f.write('    distances = LUT_DISTANCES\n')
            f.write('    \n')
            f.write('    # Handle edge cases\n')
            f.write('    if sensor_distance <= min(distances):\n')
            f.write('        idx = distances.index(min(distances))\n')
            f.write('        return positions[idx]\n')
            f.write('    if sensor_distance >= max(distances):\n')
            f.write('        idx = distances.index(max(distances))\n')
            f.write('        return positions[idx]\n')
            f.write('    \n')
            f.write('    # Linear interpolation\n')
            f.write('    for i in range(len(distances) - 1):\n')
            f.write('        d1, d2 = distances[i], distances[i + 1]\n')
            f.write('        if (d1 <= sensor_distance <= d2) or (d2 <= sensor_distance <= d1):\n')
            f.write('            p1, p2 = positions[i], positions[i + 1]\n')
            f.write('            if d2 != d1:\n')
            f.write('                return p1 + (p2 - p1) * (sensor_distance - d1) / (d2 - d1)\n')
            f.write('    \n')
            f.write('    return None\n\n')
            
            f.write('def get_sensor_error(sensor_distance):\n')
            f.write('    """\n')
            f.write('    Get the error (distance - true_position) for a sensor reading.\n')
            f.write('    \n')
            f.write('    Args:\n')
            f.write('        sensor_distance: Distance measured by the sensor (mm)\n')
            f.write('    \n')
            f.write('    Returns:\n')
            f.write('        Error in mm (positive = sensor reads too high)\n')
            f.write('    """\n')
            f.write('    true_pos = get_true_position(sensor_distance)\n')
            f.write('    if true_pos is not None:\n')
            f.write('        return sensor_distance - true_pos\n')
            f.write('    return 0\n\n')
            
            f.write('if __name__ == "__main__":\n')
            f.write('    # Test the lookup table\n')
            f.write('    print("Sensor Distance Lookup Table Test")\n')
            f.write('    print("=" * 50)\n')
            f.write('    print(f"Table has {len(LUT_POSITIONS)} entries")\n')
            f.write('    print(f"Range: {min(LUT_DISTANCES):.2f}mm to {max(LUT_DISTANCES):.2f}mm\\n")\n')
            f.write('    \n')
            f.write('    # Test some values\n')
            f.write('    test_distances = [50, 100, 150, 200, 250]\n')
            f.write('    for dist in test_distances:\n')
            f.write('        true_pos = get_true_position(dist)\n')
            f.write('        if true_pos:\n')
            f.write('            error = get_sensor_error(dist)\n')
            f.write('            print(f"Sensor: {dist:6.2f}mm → True: {true_pos:6.2f}mm | Error: {error:+6.2f}mm")\n')
    
    def write_c_header(self, filepath, lut=None):
        """Write lookup table to C header file"""
        if lut is None:
            lut = self.current_lut
        
        with open(filepath, 'w', encoding='utf-8') as f:
            f.write(f"// Sensor Distance Lookup Table: {lut.name}\n")
            f.write(f"// Generated: {datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n")
            f.write(f"// Bin Size: {lut.metadata.get('bin_size', 'N/A')} mm\n")
            f.write(f"// Method: {lut.metadata.get('method', 'N/A')}\n\n")
            
            guard = lut.name.upper().replace(' ', '_') + "_H"
            f.write(f"#ifndef {guard}\n")
            f.write(f"#define {guard}\n\n")
            
            f.write(f"#define LUT_SIZE {len(lut.compiled_positions)}\n\n")
            
            f.write("static const float lut_positions[LUT_SIZE] = {\n")
            for i, pos in enumerate(lut.compiled_positions):
                comma = "," if i < len(lut.compiled_positions) - 1 else ""
                f.write(f"    {pos:.2f}f{comma}\n")
            f.write("};\n\n")
            
            f.write("static const float lut_distances[LUT_SIZE] = {\n")
            for i, dist in enumerate(lut.compiled_distances):
                comma = "," if i < len(lut.compiled_distances) - 1 else ""
                f.write(f"    {dist:.2f}f{comma}\n")
            f.write("};\n\n")
            
            f.write(f"#endif // {guard}\n")
    
    def test_lookup_dialog(self):
        """Open test lookup dialog"""
        self.right_notebook.select(2)  # Switch to test tab
    
    def test_single_lookup(self):
        """Test a single lookup value: sensor distance -> true position"""
        if not self.current_lut:
            messagebox.showwarning("Warning", "No lookup table selected")
            return
        
        try:
            sensor_distance = float(self.test_input_var.get())
        except ValueError:
            messagebox.showerror("Error", "Invalid sensor distance value")
            return
        
        # Get selected correction method
        correction_method = self.correction_method_var.get()
        
        # Get method parameters from UI
        method_params = {}
        try:
            method_params['degree'] = int(self.poly_degree_var.get())
            method_params['k'] = int(self.k_neighbors_var.get())
            method_params['power'] = float(self.distance_power_var.get())
        except:
            pass  # Use defaults if parsing fails
        
        # Apply correction
        true_position = self.current_lut.apply_correction(sensor_distance, method=correction_method, **method_params)
        
        # Also show comparison with other method
        if correction_method == 'reverse_lookup':
            alt_position = self.current_lut.apply_correction(sensor_distance, method='error_subtraction', **method_params)
            alt_method_name = 'Error Subtraction'
        else:
            alt_position = self.current_lut.apply_correction(sensor_distance, method='reverse_lookup', **method_params)
            alt_method_name = 'Reverse Lookup'
        
        if true_position is not None:
            error = sensor_distance - true_position
            method_display = 'Reverse Lookup' if correction_method == 'reverse_lookup' else 'Error Subtraction'
            
            result_text = f"Result ({method_display}): Sensor {sensor_distance:.2f}mm → True Position {true_position:.2f}mm | Correction: {-error:.2f}mm\n"
            if alt_position is not None:
                alt_error = sensor_distance - alt_position
                diff = abs(true_position - alt_position)
                result_text += f"({alt_method_name}: {alt_position:.2f}mm | Correction: {-alt_error:.2f}mm | Diff: {diff:.3f}mm)"
            
            self.test_result_label.config(text=result_text)
        else:
            self.test_result_label.config(text="Result: Distance out of range")
    
    def batch_correction_dialog(self):
        """Open batch correction dialog"""
        self.right_notebook.select(2)  # Switch to test tab
    
    def load_batch_data(self):
        """Load CSV data for batch correction"""
        filepath = filedialog.askopenfilename(
            title="Load CSV Data",
            filetypes=[("CSV files", "*.csv"), ("All files", "*.*")]
        )
        if filepath:
            try:
                self.batch_data = []
                with open(filepath, 'r', encoding='utf-8') as f:
                    reader = csv.reader(f)
                    for row in reader:
                        self.batch_data.append(row)
                
                # Display preview
                self.batch_text.delete(1.0, tk.END)
                self.batch_text.insert(tk.END, f"Loaded {len(self.batch_data)} rows\n\n")
                for i, row in enumerate(self.batch_data[:10]):
                    self.batch_text.insert(tk.END, f"{i}: {row}\n")
                if len(self.batch_data) > 10:
                    self.batch_text.insert(tk.END, f"... and {len(self.batch_data) - 10} more rows\n")
                    
            except Exception as e:
                messagebox.showerror("Error", f"Failed to load: {str(e)}")
    
    def apply_batch_correction(self):
        """Apply lookup table correction to batch data"""
        if not self.batch_data:
            messagebox.showwarning("Warning", "No batch data loaded")
            return
        
        if not self.current_lut:
            messagebox.showwarning("Warning", "No lookup table selected")
            return
        
        self.corrected_data = []
        for row in self.batch_data:
            new_row = list(row)
            try:
                if len(row) >= 1:
                    sensor_distance = float(row[0])
                    corrected = self.current_lut.reverse_lookup(sensor_distance)
                    if corrected is not None:
                        new_row.append(f"{corrected:.2f}")
                    else:
                        new_row.append("N/A")
            except ValueError:
                new_row.append("N/A")
            self.corrected_data.append(new_row)
        
        # Display preview
        self.batch_text.delete(1.0, tk.END)
        self.batch_text.insert(tk.END, f"Corrected {len(self.corrected_data)} rows\n")
        self.batch_text.insert(tk.END, "(Last column is corrected value)\n\n")
        for i, row in enumerate(self.corrected_data[:10]):
            self.batch_text.insert(tk.END, f"{i}: {row}\n")
        if len(self.corrected_data) > 10:
            self.batch_text.insert(tk.END, f"... and {len(self.corrected_data) - 10} more rows\n")
        
        messagebox.showinfo("Success", f"Applied correction to {len(self.corrected_data)} rows")
    
    def save_corrected_data(self):
        """Save corrected data to CSV"""
        if not self.corrected_data:
            messagebox.showwarning("Warning", "No corrected data to save")
            return
        
        filepath = filedialog.asksaveasfilename(
            title="Save Corrected Data",
            defaultextension=".csv",
            filetypes=[("CSV files", "*.csv"), ("All files", "*.*")]
        )
        if filepath:
            try:
                with open(filepath, 'w', newline='', encoding='utf-8') as f:
                    writer = csv.writer(f)
                    writer.writerows(self.corrected_data)
                messagebox.showinfo("Success", f"Saved: {filepath}")
            except Exception as e:
                messagebox.showerror("Error", f"Failed to save: {str(e)}")
    
    def load_tds_for_correction(self):
        """Load a TDS Excel file for applying correction factor"""
        filepath = filedialog.askopenfilename(
            title="Load TDS File",
            filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")]
        )
        if filepath:
            try:
                wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
                
                # Try to find RAW_DATA sheet first, otherwise use active sheet
                if 'RAW_DATA' in wb.sheetnames:
                    ws = wb['RAW_DATA']
                else:
                    ws = wb.active
                
                self.tds_data = {
                    'filepath': filepath,
                    'distances': [],
                    'temperatures': [],
                    'positions': [],
                    'deltas': []
                }
                
                row_count = 0
                for row in ws.iter_rows(min_row=1, values_only=True):
                    if row and len(row) >= 4:
                        try:
                            # Excel format: distance, temp, position, delta, ...
                            distance = float(row[0]) if row[0] is not None else None
                            temp = float(row[1]) if row[1] is not None else None
                            position = float(row[2]) if row[2] is not None else None
                            delta = float(row[3]) if row[3] is not None else None
                            
                            if distance is not None and position is not None:
                                self.tds_data['distances'].append(distance)
                                self.tds_data['temperatures'].append(temp)
                                self.tds_data['positions'].append(position)
                                self.tds_data['deltas'].append(delta)
                                row_count += 1
                        except (ValueError, TypeError):
                            continue
                
                wb.close()
                
                filename = os.path.basename(filepath)
                self.tds_file_label.config(text=f"{filename} ({row_count} points)")
                self.correction_stats_label.config(text="")
                
                messagebox.showinfo("Success", f"Loaded {row_count} data points from {filename}")
                
            except Exception as e:
                messagebox.showerror("Error", f"Failed to load TDS file: {str(e)}")
    
    def update_active_lut_label(self):
        """Update the active LUT label in the test tab"""
        if self.current_lut and hasattr(self.current_lut, 'compiled_positions'):
            self.active_lut_label.config(text=self.current_lut.name, foreground="green")
        else:
            self.active_lut_label.config(text="None selected", foreground="red")
    
    def apply_lut_correction(self):
        """Apply lookup table correction to loaded TDS data"""
        if not self.tds_data:
            messagebox.showwarning("Warning", "No TDS file loaded")
            return
        
        if not self.current_lut or not hasattr(self.current_lut, 'compiled_positions'):
            messagebox.showwarning("Warning", "No lookup table selected. Please select a compiled lookup table first.")
            return
        
        # Apply lookup table correction to distances
        self.corrected_tds_data = {
            'filepath': self.tds_data['filepath'],
            'original_distances': self.tds_data['distances'].copy(),
            'corrected_positions': [],
            'temperatures': self.tds_data['temperatures'].copy(),
            'positions': self.tds_data['positions'].copy(),
            'deltas': self.tds_data['deltas'].copy(),
            'lut_name': self.current_lut.name,
            'out_of_range_count': 0
        }
        
        # Get selected correction method
        correction_method = self.correction_method_var.get()
        
        # Get method parameters from UI
        method_params = {}
        try:
            method_params['degree'] = int(self.poly_degree_var.get())
        except:
            method_params['degree'] = 3
        
        try:
            method_params['k'] = int(self.k_neighbors_var.get())
        except:
            method_params['k'] = 3
        
        try:
            method_params['power'] = float(self.distance_power_var.get())
        except:
            method_params['power'] = 1.0
        
        # Apply correction using selected method with parameters
        for distance in self.tds_data['distances']:
            corrected_position = self.current_lut.apply_correction(distance, method=correction_method, **method_params)
            if corrected_position is not None:
                self.corrected_tds_data['corrected_positions'].append(corrected_position)
            else:
                # Out of range - use original distance
                self.corrected_tds_data['corrected_positions'].append(distance)
                self.corrected_tds_data['out_of_range_count'] += 1
        
        # Calculate new deltas (corrected_position - reference_position)
        self.corrected_tds_data['corrected_deltas'] = [
            self.corrected_tds_data['corrected_positions'][i] - self.corrected_tds_data['positions'][i]
            for i in range(len(self.corrected_tds_data['positions']))
        ]
        
        # Calculate statistics
        original_errors = [abs(d) for d in self.tds_data['deltas'] if d is not None]
        corrected_errors = [abs(d) for d in self.corrected_tds_data['corrected_deltas']]
        
        # For mean, use absolute values
        orig_mean = np.mean(original_errors) if original_errors else 0
        orig_std = np.std(original_errors) if original_errors else 0
        
        corr_mean = np.mean(corrected_errors) if corrected_errors else 0
        corr_std = np.std(corrected_errors) if corrected_errors else 0
        
        # For min/max, use raw deltas to show actual range (including negative)
        orig_deltas_list = [d for d in self.tds_data['deltas'] if d is not None]
        corr_deltas_list = self.corrected_tds_data['corrected_deltas']
        
        orig_min = min(orig_deltas_list) if orig_deltas_list else 0
        orig_max = max(orig_deltas_list) if orig_deltas_list else 0
        corr_min = min(corr_deltas_list) if corr_deltas_list else 0
        corr_max = max(corr_deltas_list) if corr_deltas_list else 0
        
        improvement = ((orig_mean - corr_mean) / orig_mean * 100) if orig_mean > 0 else 0
        
        # Get method name for display with parameters
        method_display = {
            'reverse_lookup': 'Reverse Lookup (Linear)',
            'error_subtraction': 'Error Subtraction (Linear)',
            'offset_correction': 'Offset Correction',
            'spline': 'Cubic Spline',
            'nearest': 'Nearest Neighbor'
        }.get(correction_method, correction_method)
        
        # Add parameters if applicable
        if correction_method == 'polynomial':
            method_display = f"Polynomial Fit (degree={method_params.get('degree', 3)})"
        elif correction_method == 'weighted':
            k = method_params.get('k', 3)
            power = method_params.get('power', 1.0)
            method_display = f"Weighted Average (k={k}, power={power:.1f})"
        
        stats_text = (
            f"Lookup Table: {self.current_lut.name}\n"
            f"Method: {method_display}\n"
            f"Data Points: {len(self.corrected_tds_data['positions'])}\n"
            f"Out of Range: {self.corrected_tds_data['out_of_range_count']}\n\n"
            f"Original Error:\n"
            f"  Mean: {orig_mean:.3f}mm, Std: {orig_std:.3f}mm\n"
            f"  Min: {orig_min:.3f}mm, Max: {orig_max:.3f}mm\n\n"
            f"Corrected Error:\n"
            f"  Mean: {corr_mean:.3f}mm, Std: {corr_std:.3f}mm\n"
            f"  Min: {corr_min:.3f}mm, Max: {corr_max:.3f}mm\n\n"
            f"Improvement: {improvement:.1f}%"
        )
        self.correction_stats_label.config(text=stats_text)
        
        # Store correction method and parameters used
        self.corrected_tds_data['correction_method'] = correction_method
        self.corrected_tds_data['correction_params'] = method_params.copy()
        
        messagebox.showinfo("Success", 
            f"Applied LUT correction: {self.current_lut.name}\n\n"
            f"Mean error: {orig_mean:.3f}mm → {corr_mean:.3f}mm\n"
            f"Min/Max: {orig_min:.3f}/{orig_max:.3f}mm → {corr_min:.3f}/{corr_max:.3f}mm\n"
            f"Improvement: {improvement:.1f}%\n\n"
            f"{self.corrected_tds_data['out_of_range_count']} points out of LUT range")
    
    def save_corrected_tds(self):
        """Save corrected TDS data to Excel file"""
        if not self.corrected_tds_data:
            messagebox.showwarning("Warning", "No corrected TDS data to save")
            return
        
        original_name = os.path.basename(self.corrected_tds_data['filepath'])
        base_name = os.path.splitext(original_name)[0]
        lut_name = self.corrected_tds_data.get('lut_name', 'LUT').replace(' ', '_')
        default_name = f"{base_name}_corrected_{lut_name}.xlsx"
        
        filepath = filedialog.asksaveasfilename(
            title="Save Corrected TDS Data",
            defaultextension=".xlsx",
            initialfile=default_name,
            filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")]
        )
        if filepath:
            try:
                wb = openpyxl.Workbook()
                ws = wb.active
                ws.title = "RAW_DATA"
                
                # Write header
                ws.append([
                    'Reference Position (mm)', 
                    'Original Sensor Distance (mm)',
                    'LUT Corrected Position (mm)', 
                    'Temperature (C)', 
                    'Original Error (mm)',
                    'Corrected Error (mm)'
                ])
                
                # Write data
                for i in range(len(self.corrected_tds_data['positions'])):
                    ws.append([
                        self.corrected_tds_data['positions'][i],
                        self.corrected_tds_data['original_distances'][i],
                        self.corrected_tds_data['corrected_positions'][i],
                        self.corrected_tds_data['temperatures'][i],
                        self.corrected_tds_data['deltas'][i],
                        self.corrected_tds_data['corrected_deltas'][i]
                    ])
                
                # Add metadata sheet
                meta_ws = wb.create_sheet("METADATA")
                meta_ws.append(['Lookup Table', self.corrected_tds_data.get('lut_name', 'N/A')])
                meta_ws.append(['Correction Method', self.corrected_tds_data.get('correction_method', 'N/A')])
                meta_ws.append(['Original File', os.path.basename(self.corrected_tds_data['filepath'])])
                meta_ws.append(['Corrected Date', datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')])
                meta_ws.append(['Data Points', len(self.corrected_tds_data['positions'])])
                meta_ws.append(['Out of Range Points', self.corrected_tds_data.get('out_of_range_count', 0)])
                
                wb.save(filepath)
                messagebox.showinfo("Success", f"Saved corrected data to:\n{filepath}")
            except Exception as e:
                messagebox.showerror("Error", f"Failed to save: {str(e)}")
    
    def compare_correction_methods(self):
        """Compare different correction methods on current TDS data or generate test data"""
        if not self.current_lut or not hasattr(self.current_lut, 'compiled_positions'):
            messagebox.showwarning("Warning", "No compiled lookup table selected")
            return
        
        # Create comparison window
        compare_window = tk.Toplevel(self.root)
        compare_window.title("Correction Method Comparison")
        compare_window.geometry("1200x800")
        
        # Generate test distances
        lut = self.current_lut
        min_dist = min(lut.compiled_distances)
        max_dist = max(lut.compiled_distances)
        test_distances = np.linspace(min_dist, max_dist, 100)
        
        # Get current parameter values from UI
        try:
            poly_degree = int(self.poly_degree_var.get())
            k_neighbors = int(self.k_neighbors_var.get())
            distance_power = float(self.distance_power_var.get())
        except:
            poly_degree = 3
            k_neighbors = 3
            distance_power = 1.0
        
        # Apply all methods with current parameters
        methods = {
            'Reverse Lookup': ('reverse_lookup', {}),
            'Error Subtraction': ('error_subtraction', {}),
            'Cubic Spline': ('spline', {}),
            f'Polynomial (deg={poly_degree})': ('polynomial', {'degree': poly_degree}),
            f'Weighted (k={k_neighbors}, p={distance_power:.1f})': ('weighted', {'k': k_neighbors, 'power': distance_power}),
            'Nearest': ('nearest', {})
        }
        
        results = {}
        for name, (method_key, params) in methods.items():
            try:
                results[name] = [lut.apply_correction(d, method_key, **params) for d in test_distances]
            except Exception as e:
                print(f"Method {name} failed: {e}")
                results[name] = None
        
        # Remove failed methods
        results = {k: v for k, v in results.items() if v is not None}
        
        # Use reverse lookup as baseline
        baseline = results.get('Reverse Lookup', list(test_distances))
        differences = {}
        for name, vals in results.items():
            if name != 'Reverse Lookup' and vals:
                differences[name] = [abs(vals[i] - baseline[i]) if vals[i] and baseline[i] else 0
                                    for i in range(len(test_distances))]
        
        # Create plots
        fig, axes = plt.subplots(2, 2, figsize=(14, 10))
        
        # Plot 1: Corrected positions comparison
        ax1 = axes[0, 0]
        colors = ['b', 'r', 'g', 'orange', 'purple', 'brown']
        linestyles = ['-', '--', '-.', ':', '-', '--']
        for i, (name, vals) in enumerate(results.items()):
            ax1.plot(test_distances, vals, 
                    color=colors[i % len(colors)], 
                    linestyle=linestyles[i % len(linestyles)],
                    label=name, linewidth=1.5, alpha=0.8)
        ax1.plot([min_dist, max_dist], [min_dist, max_dist], 'k:', label='Ideal (y=x)', alpha=0.3, linewidth=1)
        ax1.set_xlabel('Sensor Distance (mm)')
        ax1.set_ylabel('Corrected Position (mm)')
        ax1.set_title('All Correction Methods Comparison')
        ax1.legend(fontsize=8, loc='best')
        ax1.grid(True, alpha=0.3)
        
        # Plot 2: Difference from baseline (Reverse Lookup)
        ax2 = axes[0, 1]
        for i, (name, diffs) in enumerate(differences.items()):
            if diffs:
                ax2.plot(test_distances, diffs, 
                        color=colors[(i+1) % len(colors)], 
                        label=name, linewidth=1.5, alpha=0.8)
        ax2.set_xlabel('Sensor Distance (mm)')
        ax2.set_ylabel('Difference from Reverse Lookup (mm)')
        max_diff = max([max(d) for d in differences.values()]) if differences else 0
        mean_diff = np.mean([np.mean(d) for d in differences.values()]) if differences else 0
        ax2.set_title(f'Difference from Baseline\nMax: {max_diff:.4f}mm, Avg: {mean_diff:.4f}mm')
        ax2.legend(fontsize=8, loc='best')
        ax2.grid(True, alpha=0.3)
        
        # Plot 3: Correction amount for each method
        ax3 = axes[1, 0]
        for i, (name, vals) in enumerate(results.items()):
            corrections = [test_distances[j] - vals[j] if vals[j] else 0 
                          for j in range(len(test_distances))]
            ax3.plot(test_distances, corrections, 
                    color=colors[i % len(colors)],
                    linestyle=linestyles[i % len(linestyles)],
                    label=name, linewidth=1.5, alpha=0.8)
        ax3.axhline(y=0, color='k', linestyle='-', alpha=0.3)
        ax3.set_xlabel('Sensor Distance (mm)')
        ax3.set_ylabel('Correction Applied (mm)')
        ax3.set_title('Correction Amount by Method')
        ax3.legend(fontsize=8, loc='best')
        ax3.grid(True, alpha=0.3)
        
        # Plot 4: Lookup table visualization
        ax4 = axes[1, 1]
        ax4.plot(lut.compiled_positions, lut.compiled_distances, 'o-', markersize=6, label='LUT Data')
        ax4.plot([min(lut.compiled_positions), max(lut.compiled_positions)], 
                [min(lut.compiled_positions), max(lut.compiled_positions)], 
                'r--', alpha=0.7, label='Ideal (y=x)')
        ax4.set_xlabel('Position (mm)')
        ax4.set_ylabel('Distance (mm)')
        ax4.set_title(f'Lookup Table: {lut.name}')
        ax4.legend()
        ax4.grid(True, alpha=0.3)
        
        fig.tight_layout()
        
        canvas = FigureCanvasTkAgg(fig, compare_window)
        canvas.get_tk_widget().pack(fill=tk.BOTH, expand=True)
        
        toolbar_frame = ttk.Frame(compare_window)
        toolbar_frame.pack(fill=tk.X)
        NavigationToolbar2Tk(canvas, toolbar_frame)
        
        # Add info text
        info_frame = ttk.Frame(compare_window)
        info_frame.pack(fill=tk.X, padx=10, pady=5)
        
        # Build summary text
        summary_lines = ["Method Comparison Summary:\n"]
        summary_lines.append(f"Parameters: Polynomial degree={poly_degree}, Weighted k={k_neighbors}, power={distance_power:.1f}")
        summary_lines.append("LINEAR: Reverse Lookup (fast, standard), Error Subtraction (alternative linear)")
        summary_lines.append("SMOOTH: Spline (smooth curves), Polynomial (adjustable degree), Weighted (adjustable k & power)")
        summary_lines.append("SIMPLE: Nearest Neighbor (no interpolation)\n")
        
        if differences:
            max_diff = max([max(d) for d in differences.values()])
            avg_diff = np.mean([np.mean(d) for d in differences.values()])
            summary_lines.append(f"Max difference from baseline: {max_diff:.4f}mm")
            summary_lines.append(f"Avg difference from baseline: {avg_diff:.4f}mm\n")
        
        summary_lines.append("Recommendation: Use 'Reverse Lookup' for most cases (fast, accurate).")
        summary_lines.append("Use 'Spline' or 'Polynomial' if you need smoother corrections.")
        
        info_text = "\n".join(summary_lines)
        
        info_label = ttk.Label(info_frame, text=info_text, justify=tk.LEFT)
        info_label.pack()
    
    def run_optimization(self):
        """Run optimization to find best correction method and parameters"""
        if not self.tds_data:
            messagebox.showerror("Error", "Please load TDS data in the 'Test & Apply' tab first.")
            return
        
        if not self.current_lut or not hasattr(self.current_lut, 'compiled_positions'):
            messagebox.showerror("Error", "No compiled lookup table available. Please create one first.")
            return
        
        # Parse search parameters
        try:
            poly_min = int(self.opt_poly_min_var.get())
            poly_max = int(self.opt_poly_max_var.get())
            k_min = int(self.opt_k_min_var.get())
            k_max = int(self.opt_k_max_var.get())
            power_values = [float(x.strip()) for x in self.opt_power_values_var.get().split(',')]
        except ValueError:
            messagebox.showerror("Error", "Invalid parameter values. Please check your inputs.")
            return
        
        # Build test configurations
        configs = []
        
        # Linear methods (no parameters)
        configs.append(('Reverse Lookup', 'reverse_lookup', {}))
        configs.append(('Error Subtraction', 'error_subtraction', {}))
        configs.append(('Nearest Neighbor', 'nearest', {}))
        
        # Spline (no parameters)
        configs.append(('Cubic Spline', 'spline', {}))
        
        # Polynomial with different degrees
        for degree in range(poly_min, poly_max + 1):
            configs.append((f'Polynomial (deg={degree})', 'polynomial', {'degree': degree}))
        
        # Weighted average with different k and power combinations
        for k in range(k_min, k_max + 1):
            for power in power_values:
                configs.append((f'Weighted (k={k}, p={power:.1f})', 'weighted', {'k': k, 'power': power}))
        
        # Clear previous results
        for item in self.opt_tree.get_children():
            self.opt_tree.delete(item)
        
        self.opt_progress_label.config(text=f"Testing {len(configs)} configurations...")
        self.root.update()
        
        # Calculate baseline (original) error
        baseline_errors = [abs(d) for d in self.tds_data['deltas'] if d is not None]
        baseline_mean = np.mean(baseline_errors)
        
        # Test each configuration
        results = []
        for i, (name, method, params) in enumerate(configs):
            try:
                # Apply correction to all data points
                corrected_positions = []
                for distance in self.tds_data['distances']:
                    try:
                        pos = self.current_lut.apply_correction(distance, method=method, **params)
                        corrected_positions.append(pos)
                    except:
                        corrected_positions.append(distance)  # Fallback
                
                # Calculate errors
                corrected_deltas = [
                    corrected_positions[j] - self.tds_data['positions'][j]
                    for j in range(len(self.tds_data['positions']))
                ]
                
                abs_errors = [abs(d) for d in corrected_deltas]
                mean_error = np.mean(abs_errors)
                std_error = np.std(abs_errors)
                improvement = ((baseline_mean - mean_error) / baseline_mean * 100) if baseline_mean > 0 else 0
                
                results.append({
                    'name': name,
                    'method': method,
                    'params': params.copy(),
                    'mean_error': mean_error,
                    'std_error': std_error,
                    'improvement': improvement,
                    'corrected_positions': corrected_positions,
                    'corrected_deltas': corrected_deltas
                })
                
            except Exception as e:
                # Skip failed configurations
                pass
            
            # Update progress
            if (i + 1) % 10 == 0 or (i + 1) == len(configs):
                self.opt_progress_label.config(text=f"Tested {i + 1}/{len(configs)} configurations...")
                self.root.update()
        
        # Sort by mean error (lowest is best)
        results.sort(key=lambda x: x['mean_error'])
        
        # Store results
        self.optimization_results = results
        
        # Display top 20 results
        for i, result in enumerate(results[:20], 1):
            param_str = ', '.join([f"{k}={v}" for k, v in result['params'].items()]) if result['params'] else '-'
            
            self.opt_tree.insert('', 'end', values=(
                i,
                result['name'],
                param_str,
                f"{result['mean_error']:.4f}",
                f"{result['improvement']:.2f}",
                f"{result['std_error']:.4f}"
            ))
        
        # Update progress label with summary
        if results:
            best = results[0]
            self.opt_progress_label.config(
                text=f"Complete! Best: {best['name']} - {best['mean_error']:.4f}mm ({best['improvement']:.1f}% improvement)",
                foreground='green'
            )
            
            messagebox.showinfo(
                "Optimization Complete",
                f"Tested {len(configs)} configurations.\n\n"
                f"Best Method: {best['name']}\n"
                f"Mean Error: {best['mean_error']:.4f} mm\n"
                f"Improvement: {best['improvement']:.1f}%\n\n"
                f"See results table for details."
            )
        else:
            self.opt_progress_label.config(text="No valid results found.", foreground='red')
    
    def apply_best_method(self):
        """Apply the best method found by optimization to the current TDS data"""
        if not self.optimization_results:
            messagebox.showerror("Error", "No optimization results available. Run optimization first.")
            return
        
        best = self.optimization_results[0]
        
        # Apply the best configuration
        self.corrected_tds_data = {
            'filepath': self.tds_data['filepath'],
            'original_distances': self.tds_data['distances'].copy(),
            'corrected_positions': best['corrected_positions'].copy(),
            'temperatures': self.tds_data['temperatures'].copy(),
            'positions': self.tds_data['positions'].copy(),
            'deltas': self.tds_data['deltas'].copy(),
            'corrected_deltas': best['corrected_deltas'].copy(),
            'lut_name': self.current_lut.name,
            'correction_method': best['method'],
            'correction_params': best['params'].copy(),
            'out_of_range_count': 0
        }
        
        # Update stats display in Test & Apply tab
        baseline_errors = [abs(d) for d in self.tds_data['deltas'] if d is not None]
        baseline_mean = np.mean(baseline_errors) if baseline_errors else 0
        
        stats_text = (
            f"Applied OPTIMIZED Method:\n"
            f"Method: {best['name']}\n"
            f"Mean Error: {baseline_mean:.4f}mm → {best['mean_error']:.4f}mm\n"
            f"Improvement: {best['improvement']:.1f}%\n"
            f"Std Dev: {best['std_error']:.4f}mm\n\n"
            f"This was automatically selected as the best configuration.\n"
            f"You can now save the corrected data in the 'Test & Apply' tab."
        )
        self.correction_stats_label.config(text=stats_text)
        
        messagebox.showinfo(
            "Applied Best Method",
            f"Applied: {best['name']}\n\n"
            f"Mean Error: {best['mean_error']:.4f} mm\n"
            f"Improvement: {best['improvement']:.1f}%\n\n"
            f"Switch to 'Test & Apply' tab to save or visualize."
        )
        
        # Switch to Test & Apply tab
        self.right_notebook.select(2)
    
    def visualize_top_methods(self):
        """Visualize the top 5 methods from optimization"""
        if not self.optimization_results:
            messagebox.showerror("Error", "No optimization results available. Run optimization first.")
            return
        
        # Create visualization window
        viz_window = tk.Toplevel(self.root)
        viz_window.title("Top 5 Optimization Results")
        viz_window.geometry("1200x800")
        
        fig, axes = plt.subplots(2, 2, figsize=(14, 10))
        
        top_5 = self.optimization_results[:5]
        colors = ['blue', 'red', 'green', 'orange', 'purple']
        
        positions = self.tds_data['positions']
        
        # Plot 1: Corrected positions comparison
        ax1 = axes[0, 0]
        for i, result in enumerate(top_5):
            ax1.scatter(positions, result['corrected_positions'], alpha=0.5, s=10, 
                       color=colors[i], label=f"#{i+1}: {result['name']}")
        min_val = min(positions)
        max_val = max(positions)
        ax1.plot([min_val, max_val], [min_val, max_val], 'k--', alpha=0.3, label='Ideal')
        ax1.set_xlabel('Reference Position (mm)')
        ax1.set_ylabel('Corrected Position (mm)')
        ax1.set_title('Top 5 Methods - Corrected Positions')
        ax1.legend(fontsize=8, loc='best')
        ax1.grid(True, alpha=0.3)
        
        # Plot 2: Error comparison
        ax2 = axes[0, 1]
        for i, result in enumerate(top_5):
            ax2.scatter(positions, result['corrected_deltas'], alpha=0.5, s=10, 
                       color=colors[i], label=f"#{i+1}: {result['name']}")
        ax2.axhline(y=0, color='k', linestyle='--', alpha=0.3)
        ax2.set_xlabel('Reference Position (mm)')
        ax2.set_ylabel('Error (mm)')
        ax2.set_title('Top 5 Methods - Error Distribution')
        ax2.legend(fontsize=8, loc='best')
        ax2.grid(True, alpha=0.3)
        
        # Plot 3: Mean error comparison bar chart
        ax3 = axes[1, 0]
        names = [f"#{i+1}\n{r['name'][:20]}" for i, r in enumerate(top_5)]
        mean_errors = [r['mean_error'] for r in top_5]
        bars = ax3.bar(range(len(top_5)), mean_errors, color=colors)
        ax3.set_xticks(range(len(top_5)))
        ax3.set_xticklabels(names, fontsize=8)
        ax3.set_ylabel('Mean Absolute Error (mm)')
        ax3.set_title('Top 5 Methods - Mean Error Comparison')
        ax3.grid(True, alpha=0.3, axis='y')
        
        # Add value labels on bars
        for i, (bar, val) in enumerate(zip(bars, mean_errors)):
            ax3.text(bar.get_x() + bar.get_width()/2, val, f'{val:.4f}',
                    ha='center', va='bottom', fontsize=8)
        
        # Plot 4: Statistics summary
        ax4 = axes[1, 1]
        ax4.axis('off')
        
        summary_text = "Top 5 Results Summary:\n\n"
        for i, result in enumerate(top_5, 1):
            param_str = ', '.join([f"{k}={v}" for k, v in result['params'].items()]) if result['params'] else 'none'
            summary_text += f"#{i}: {result['name']}\n"
            summary_text += f"   Params: {param_str}\n"
            summary_text += f"   Mean Error: {result['mean_error']:.4f}mm\n"
            summary_text += f"   Std Dev: {result['std_error']:.4f}mm\n"
            summary_text += f"   Improvement: {result['improvement']:.1f}%\n\n"
        
        ax4.text(0.1, 0.9, summary_text, transform=ax4.transAxes,
                fontsize=9, verticalalignment='top', family='monospace')
        
        fig.tight_layout()
        
        canvas = FigureCanvasTkAgg(fig, viz_window)
        canvas.get_tk_widget().pack(fill=tk.BOTH, expand=True)
        
        toolbar_frame = ttk.Frame(viz_window)
        toolbar_frame.pack(fill=tk.X)
        NavigationToolbar2Tk(canvas, toolbar_frame)
    
    def export_optimization_results(self):
        """Export optimization results to CSV file"""
        if not self.optimization_results:
            messagebox.showerror("Error", "No optimization results available. Run optimization first.")
            return
        
        filepath = filedialog.asksaveasfilename(
            title="Export Optimization Results",
            defaultextension=".csv",
            initialfile="optimization_results.csv",
            filetypes=[("CSV files", "*.csv"), ("All files", "*.*")]
        )
        
        if filepath:
            try:
                with open(filepath, 'w', newline='', encoding='utf-8') as f:
                    import csv
                    writer = csv.writer(f)
                    writer.writerow(['Rank', 'Method', 'Parameters', 'Mean_Error_mm', 'Std_Error_mm', 'Improvement_%'])
                    
                    for i, result in enumerate(self.optimization_results, 1):
                        param_str = '; '.join([f"{k}={v}" for k, v in result['params'].items()]) if result['params'] else '-'
                        writer.writerow([
                            i,
                            result['name'],
                            param_str,
                            f"{result['mean_error']:.6f}",
                            f"{result['std_error']:.6f}",
                            f"{result['improvement']:.2f}"
                        ])
                
                messagebox.showinfo("Success", f"Exported {len(self.optimization_results)} results to:\n{filepath}")
            except Exception as e:
                messagebox.showerror("Error", f"Failed to export results:\n{str(e)}")
    
    def show_correction_comparison(self):
        """Show comparison plot of original vs corrected data"""
        if not self.corrected_tds_data:
            messagebox.showwarning("Warning", "No corrected data available. Apply correction factor first.")
            return
        
        # Create comparison plot window
        plot_window = tk.Toplevel(self.root)
        plot_window.title("Correction Factor Comparison")
        plot_window.geometry("1000x700")
        
        fig, axes = plt.subplots(2, 2, figsize=(12, 10))
        
        positions = self.corrected_tds_data['positions']
        orig_distances = self.corrected_tds_data['original_distances']
        corr_positions = self.corrected_tds_data['corrected_positions']
        orig_deltas = self.corrected_tds_data['deltas']
        corr_deltas = self.corrected_tds_data['corrected_deltas']
        
        # Calculate error statistics
        orig_abs_errors = [abs(d) for d in orig_deltas]
        corr_abs_errors = [abs(d) for d in corr_deltas]
        
        orig_mean = np.mean(orig_abs_errors)
        orig_min = min(orig_deltas)  # Use raw deltas for min/max to show actual range
        orig_max = max(orig_deltas)
        
        corr_mean = np.mean(corr_abs_errors)
        corr_min = min(corr_deltas)  # Use raw deltas for min/max to show actual range
        corr_max = max(corr_deltas)
        
        # Plot 1: Original Sensor Distance vs Reference Position
        ax1 = axes[0, 0]
        ax1.scatter(positions, orig_distances, alpha=0.5, s=10, label='Sensor Reading')
        min_val = min(min(positions), min(orig_distances))
        max_val = max(max(positions), max(orig_distances))
        ax1.plot([min_val, max_val], [min_val, max_val], 'r--', alpha=0.7, label='Ideal (y=x)')
        ax1.set_xlabel('Reference Position (mm)')
        ax1.set_ylabel('Sensor Distance (mm)')
        ax1.set_title('Original Sensor Data')
        ax1.grid(True, alpha=0.3)
        ax1.legend()
        
        # Plot 2: Corrected Position vs Reference Position
        ax2 = axes[0, 1]
        ax2.scatter(positions, corr_positions, alpha=0.5, s=10, color='green', label='LUT Corrected')
        min_val = min(min(positions), min(corr_positions))
        max_val = max(max(positions), max(corr_positions))
        ax2.plot([min_val, max_val], [min_val, max_val], 'r--', alpha=0.7, label='Ideal (y=x)')
        ax2.set_xlabel('Reference Position (mm)')
        ax2.set_ylabel('Corrected Position (mm)')
        ax2.set_title(f'LUT Corrected Data ({self.corrected_tds_data.get("lut_name", "N/A")})')
        ax2.grid(True, alpha=0.3)
        ax2.legend()
        
        # Plot 3: Original Error vs Position
        ax3 = axes[1, 0]
        ax3.scatter(positions, orig_deltas, alpha=0.5, s=10, color='orange')
        ax3.axhline(y=0, color='r', linestyle='--', alpha=0.7)
        ax3.set_xlabel('Position (mm)')
        ax3.set_ylabel('Error (mm)')
        ax3.set_title(f'Original Error\nMean: {orig_mean:.3f}mm | Min: {orig_min:.3f}mm | Max: {orig_max:.3f}mm')
        ax3.grid(True, alpha=0.3)
        
        # Plot 4: Corrected Error vs Position
        ax4 = axes[1, 1]
        ax4.scatter(positions, corr_deltas, alpha=0.5, s=10, color='green')
        ax4.axhline(y=0, color='r', linestyle='--', alpha=0.7)
        ax4.set_xlabel('Position (mm)')
        ax4.set_ylabel('Error (mm)')
        ax4.set_title(f'Corrected Error\nMean: {corr_mean:.3f}mm | Min: {corr_min:.3f}mm | Max: {corr_max:.3f}mm')
        ax4.grid(True, alpha=0.3)
        
        fig.tight_layout()
        
        canvas = FigureCanvasTkAgg(fig, plot_window)
        canvas.get_tk_widget().pack(fill=tk.BOTH, expand=True)
        
        toolbar_frame = ttk.Frame(plot_window)
        toolbar_frame.pack(fill=tk.X)
        NavigationToolbar2Tk(canvas, toolbar_frame)


def main():
    root = tk.Tk()
    app = LookupTableGUI(root)
    root.mainloop()


if __name__ == "__main__":
    main()
